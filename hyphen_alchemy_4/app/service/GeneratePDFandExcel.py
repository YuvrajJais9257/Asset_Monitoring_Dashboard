"""
Module: Report Generation Utilities

This module contains utility functions to generate reports in both PDF and Excel formats. 
The reports are generated based on data retrieved from MySQL or PostgreSQL databases. 
Additionally, this module provides text-processing functions to clean HTML text, 
detect long text columns, and calculate dynamic column widths for the reports.
"""
import re
from datetime import datetime
import vertica_python
from typing import List, Dict, Any
from reportlab.lib import colors
from reportlab.lib.pagesizes import A1, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as excel_img
import mysql.connector
import psycopg2
from openpyxl.utils import get_column_letter
import utilities.loggings as LOGGINGS
from utilities.config import config
import pandas as pd
import csv
custom_page_size = A1
customer_id = 1
from  psycopg2.extras import RealDictCursor
from psycopg2 import Error


def html_text_convertor(html: str) -> str:
    """
    Convert HTML text to plain text.
    
    Args:
    html (str): HTML text.
    
    Returns:
    str: Plain text.
    """
    try:
        text = re.sub(r"<br\s*/?>", " ", html)
        text = re.sub(r"<style[^>]*>.*?</style>|<script[^>]*>.*?</script>", "", text)
        text = re.sub(r"<.*?>", "", text)
        text = re.sub(r"&\w+;", " ", text)
        return text
    except re.error as err:
        raise ValueError(f"Error: {err}")

def correct_rows(row: List[Any]) -> List[Any]:
    """
    Correct rows by converting HTML text, handling None values, and formatting datetime objects.
    
    Args:
    row (List[Any]): Row data.
    
    Returns:
    List[Any]: Corrected row data.
    """
    modified_row = []
    for item in row:
        if isinstance(item, str) and ("<" in item or ">" in item):
            new_item = html_text_convertor(item)
            modified_row.append(new_item)
        elif item is None:
            modified_row.append("None")
        elif isinstance(item, list):
            modified_row.append(str(item))
        elif isinstance(item, datetime):
            modified_row.append(item.strftime("%Y-%m-%d %H:%M:%S"))
        else:
            modified_row.append(item)
    return modified_row

def split_text(text: str, max_length: int) -> List[str]:
    """
    Split text into chunks of maximum length.
    
    Args:
    text (str): Text to split.
    max_length (int): Maximum length of each chunk.
    
    Returns:
    List[str]: List of text chunks.
    """
    if not text:
        return []
    return [text[i : i + max_length] for i in range(0, len(text), max_length)]

def detect_long_text_columns(rows: List[Dict[str, Any]], max_length: int = 10) -> List[str]:
    """
    Detect columns with long text values.
    
    Args:
    rows (List[Dict[str, Any]]): List of row data.
    max_length (int): Maximum length threshold. Defaults to 10.
    
    Returns:
    List[str]: List of column names with long text values.
    """
    long_text_columns = set()
    for key in rows[0].keys():
        for row in rows:
            if isinstance(row[key], str) and len(row[key]) > max_length:
                long_text_columns.add(key)
                break
    return list(long_text_columns)

def detect_long_text_columns_postgres(rows, columns, max_length=10):
    """
    Detect long text columns in PostgreSQL result with explicit column information.
    """
    long_text_columns = []
    keys = [col.name for col in columns]
    for i, key in enumerate(keys):
        for row in rows:
            if isinstance(row[i], str) and len(row[i]) > max_length:
                long_text_columns.append(key)
                break

    return long_text_columns

def calculate_column_widths_mysql(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content.
    """
    flag = 0
    max_column_widths = [len(col) for col in rows[0].keys()]
    for row in rows:
        for i, key in enumerate(row.keys()):
            max_width = len(str(row[key]))
            max_column_widths[i] = max(max_width, max_column_widths[i])

    total_width = sum(max_column_widths)

    while total_width < page_width:
        scaling_factor = (page_width) / total_width
        max_column_widths = [
            max(min_absolute_width, (width * scaling_factor))
            for width in max_column_widths
        ]
        total_width = sum(max_column_widths)
        flag = 1

    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [
                max(min_absolute_width, width * scaling_factor)
                for width in max_column_widths
            ]

    return column_widths

def calculate_column_widths_postgres(
    page_width, rows, min_absolute_width=80
):
    """
    Calculate column widths dynamically based on page width, number of columns, and\
          content for PostgreSQL cursor result.
    """
    flag = 0
    max_column_widths = [len(str(value)) for value in rows[0]]
    for row in rows:
        for i, value in enumerate(row):
            max_width = len(str(value))
            max_column_widths[i] = max(max_width, max_column_widths[i])
    total_width = sum(max_column_widths)
    count = 0
    while total_width < page_width:
        count += 1
        scaling_factor = (page_width) / total_width
        max_column_widths = [
            max(min_absolute_width, (width * scaling_factor))
            for width in max_column_widths
        ]
        total_width = sum(max_column_widths)
        flag = 1

    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [
                max(min_absolute_width, int(width * scaling_factor))
                for width in max_column_widths
            ]

    return column_widths

def fetch_image_from_mysql(customer_id):
    """
    Fetches an image stored as a BLOB in a MySQL database and saves it to a file.

    :param db_user: MySQL username
    :param db_password: MySQL password
    :param db_host: MySQL host (e.g., 'localhost')
    :param db_name: Name of the database
    :param table_name: Table containing the image
    :param column_name: Column storing the image as a BLOB
    :param condition: SQL WHERE condition (e.g., "id=1")
    :param output_path: Path where the image should be saved
    :return: Path to the saved image file or None if no image is found
    """

    try:
        # Connect to MySQL database
        conn = mysql.connector.connect(
            host=config["mysql"]["mysql_host"],
            user=config["mysql"]["mysql_username"],
            port=config["mysql"]["mysql_port"],
            password=config["mysql"]["mysql_password"],
            database=config["mysql"]["mysql_new_schema"]
        )
        cursor = conn.cursor()

        # Fetch the BLOB data
        query = f"SELECT pdf_logo FROM img WHERE customer_id = %s LIMIT 1"
        cursor.execute(query,(customer_id,))
        result = cursor.fetchone()

        if result and result[0]:  # Ensure an image was found
            img_data = result[0]

            # Save image to file
            with open('logo.png', "wb") as img_file:
                img_file.write(img_data)

            return 'logo.png'

        else:
            return 'logo.png'

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return None

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            
def fetch_image_from_postgres(customer_id):
    """
    Fetches an image stored as a BLOB in a MySQL database and saves it to a file.

    :param db_user: MySQL username
    :param db_password: MySQL password
    :param db_host: MySQL host (e.g., 'localhost')
    :param db_name: Name of the database
    :param table_name: Table containing the image
    :param column_name: Column storing the image as a BLOB
    :param condition: SQL WHERE condition (e.g., "id=1")
    :param output_path: Path where the image should be saved
    :return: Path to the saved image file or None if no image is found
    """
    
    cursor_logger = LOGGINGS.CustomLogger()
    logging = cursor_logger.setup_logger()
    logging.info("start fetching image")

    try:
        # Connect to postgres
        conn = psycopg2.connect(
            host=config['postgres']['postgres_host'],
            user=config['postgres']['postgres_username'],
            port=config['postgres']['postgres_port'],
            password=config['postgres']['postgres_password'],
            dbname=config['postgres']['postgres_schema']
        )
        cursor = conn.cursor()

        # Fetch the BLOB data
        query = f"SELECT pdf_logo FROM img WHERE customer_id = %s LIMIT 1"
        cursor.execute(query,(customer_id,))
        result = cursor.fetchone()

        if result and result[0]:  # Ensure an image was found
            img_data = result[0]
            
            if isinstance(img_data, memoryview):
                img_data = img_data.tobytes()

            # Save image to file
            with open('logo.png', "wb") as img_file:
                img_file.write(img_data)

            return 'logo.png'

        else:
            return 'logo.png'

    except Error as err:
        logging.error("Error: %s",err)
        return None

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def first_page_content(db_type,canvas, doc, report_title, dates, custom_page_size):
    """
    Draw the first page content of the report.
    """
    dates = dates.strftime("%Y-%m-%d %H:%M:%S")
    page_width, page_height = doc.pagesize
    x_text = 350
    # y_text = custom_page_size[0] - 40
    y_text = custom_page_size[0] - 60
    report_title = report_title.upper()
    # logo_path = "logo.png"
    
    if db_type == "mysql":
        logo_path = fetch_image_from_mysql(customer_id)
    elif db_type == "postgres":
        logo_path = fetch_image_from_postgres(customer_id)

    canvas.setFont("Helvetica-Bold", 15)
    canvas.drawString(x_text, y_text, report_title)

    x_text = custom_page_size[1] - 200
    canvas.setFont("Helvetica-Bold", 15)
    canvas.drawString(x_text, y_text, dates)
    if custom_page_size == "A1":
        canvas.drawImage(logo_path, 120, page_height - 80, width=150, height=50)
    else:
        canvas.drawImage(logo_path, 50, page_height - 70, width=80, height=40)

def generate_pdf_report(report_query, filename, db_config, title_text, db_type):
    """
    Generate PDF report from database query results with standardized handling across different databases.
    
    Args:
    report_query (str): SQL query to execute
    filename (str): Output PDF filename
    db_config (dict): Database connection configuration
    title_text (str): Report title
    db_type (str): Database type ('mysql', 'postgres', or 'vertica')
    """
    try:
        
        cursor_logger = LOGGINGS.CustomLogger()
        logging = cursor_logger.setup_logger()
        logging.info("PDF report generation started.")
        

        global custom_page_size
        time = datetime.now()

        # Standardized connection handling
        if db_type == "mysql":
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor(dictionary=True)
        elif db_type == "postgres":
            conn = psycopg2.connect(**db_config)
            logging.info("Connected to PostgreSQL database.")
            logging.info(conn)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
        elif db_type == "vertica":
            conn = vertica_python.connect(**db_config)
            cursor = conn.cursor()
        else:
            raise ValueError(f"Unsupported database type: {db_type}")

        cursor.execute(report_query)
        
        # Get column names for postgres/vertica
        columns = [desc[0] for desc in cursor.description]
        logging.info("%s ",columns)
        
        batch_size = 1000
        long_text_columns = None
        col_widths = None
        all_tables = []

        # Set page size based on column count
        column_count = len(cursor.description)
        if column_count <= 5:
            custom_page_size = A4
        else:
            custom_page_size = A1
            
        if column_count > 12:
            return {
                "status": 500,
                "Error": "Column limit exceeded, max columns allowed = 12",
            }

        page_width = custom_page_size[0]

        # Process results in batches
        for batch_count, result_batch in enumerate(
            iter(lambda: cursor.fetchmany(batch_size), [])
        ):
            if not result_batch:
                break

            # Convert postgres/vertica results to dict format like mysql
            if db_type in ["postgres", "vertica"]:
                dict_batch = []
                for row in result_batch:
                    dict_row = {columns[i]: value for i, value in enumerate(row)}
                    dict_batch.append(dict_row)
                result_batch = dict_batch

            # Detect long text columns and calculate widths
            if long_text_columns is None:
                long_text_columns = detect_long_text_columns(result_batch)
                col_widths = calculate_column_widths_mysql(
                    page_width, len(result_batch[0]), result_batch
                )

            # Prepare table data
            table_data = []
            keys = list(result_batch[0].keys())
            table_data.append(tuple(keys))

            # Handle long text columns
            font_size = 8
            for col in long_text_columns:
                col_index = keys.index(col)
                max_content_length = 3000
                line_length = int(0.19 * col_widths[col_index])
                for row in result_batch:
                    row[col] = "\n".join(
                        split_text(
                            str(row[col])[:max_content_length],
                            line_length
                        )
                    )

            # Process rows
            for row in result_batch:
                new_row = list(row.values())
                modified_row = correct_rows(new_row)
                table_data.append(tuple(modified_row))

            # Create and style table
            table = Table(table_data, colWidths=col_widths)
            style = TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), font_size),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
            table.setStyle(style)
            all_tables.append(table)

        # Generate PDF
        def on_first_page_function(canvas, doc):
            first_page_content(db_type,canvas, doc, title_text, time, custom_page_size)

        output_doc = SimpleDocTemplate(
            filename, pagesize=(custom_page_size[1], custom_page_size[0])
        )
        output_doc.build(all_tables, onFirstPage=on_first_page_function)

    except Exception as e:
        logging.error(f"Error in pdf Exception: {str(e)}")
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()

def generate_excel_report(df,filename, title_text,start_date,end_date,db_type):
    try:
        cursor_logger = LOGGINGS.CustomLogger()
        logging = cursor_logger.setup_logger()
        
        df.columns = [col.replace('_', ' ').title() for col in df.columns]
        
        logging.info("excel report generation started.")
    
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
    
        # Add logo
        # logo_path = "logo.png"
        if db_type == "mysql":
            logo_path = fetch_image_from_mysql(customer_id)
        elif db_type == "postgres":
            logo_path = fetch_image_from_postgres(customer_id)
        img = excel_img(logo_path)
        logging.info("fetched img ")
        # img.width = 100
        img.width = 120
        # img.height = 50
        img.height = 40
        ws.add_image(img, "A1")  # Add image to cell A1
        ws.row_dimensions[1].height = 70
        # ws.row_dimensions[1].height = 50  # Adjust row height to accommodate the logo
       
        title_text = title_text.upper()
    
        # Add title
        if len(df.columns) >= 1:
            title_col_range = (
                f"B1:{chr(65 + len(df.columns))}1"  # Determine column range for title
            )
        else:
            title_col_range = "B1:B1"
        ws.merge_cells(title_col_range)  # Merge cells for title
        title_cell = ws["B1"]  # Assuming the title starts from column B
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=16)
    
        # Add from & to date
        from_date = f"From: {start_date}"
        to_date = f"To: {end_date}"
        ws.merge_cells("B2:C2")
        ws["B2"] = from_date
        ws.merge_cells("D2:E2")
        ws["D2"] = to_date
        ws["B2"].alignment = Alignment(horizontal="center")
        ws["D2"].alignment = Alignment(horizontal="center")
        if df.empty:
            ws.merge_cells('B4:E4')  # Merge cells for the "No data available" message
            no_data_cell = ws['B4']
            no_data_cell.value = "No data available"
            no_data_cell.alignment = Alignment(horizontal='center')
            no_data_cell.font = Font(bold=True, size=14)
        else:
            # Write DataFrame to Excel
            for r_idx, row in enumerate(df.iterrows(), start=4):
                for c_idx, value in enumerate(row[1], start=1):
                    if value is None:
                        value = str(value)
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
    
        # Apply borders to data
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
    
        max_col_index = len(df.columns)
    
        for row in ws.iter_rows(
            min_row=4, max_row=ws.max_row, min_col=1, max_col=max_col_index
            # min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = border_style
    
        # Set column names (headers)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_style
    
        # Set column widths
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            column_letter = chr(64 + col_idx)  # Convert column index to letter
            ws.column_dimensions[column_letter].width = adjusted_width
    
        # Save the workbook
        wb.save(filename)
    except Exception as e:
        logging.error("Error in excel: %s",e)

def generate_excel_report_2(df, filename, title_text):
    """
    Generate Excel report from provided DataFrame.

    Args:
    df (DataFrame): DataFrame containing the data
    filename (str): Output Excel filename
    title_text (str): Report title
    """
    try:
        # Set up logging
        cursor_logger = LOGGINGS.CustomLogger()
        logging = cursor_logger.setup_logger()
        logging.info("Excel report generation started.")

        df.columns = [col.replace('_', ' ').title() for col in df.columns]
        logging.info(f"Columns renamed: {df.columns.tolist()}")

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        logging.info("Workbook and worksheet created.")
        
        # Add logo
        logo_path = "logo.png"
        img = excel_img(logo_path)
        img.width = 100
        img.height = 50
        ws.add_image(img, "A1")  # Add image to cell A1
        ws.row_dimensions[1].height = 50  # Adjust row height to accommodate the logo
        # Add title
        if len(df.columns) >= 1:
            last_col_letter = get_column_letter(len(df.columns) + 1)
            title_col_range = f"B1:{last_col_letter}1"  # Determine column range for title
        else:
            title_col_range = "B1:B1"
        ws.merge_cells(title_col_range)  # Merge cells for title
        title_cell = ws["B1"]  # Assuming the title starts from column B
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=16)
        logging.info(f"Title added: {title_text}")

        if df.empty:
            ws.merge_cells('B4:E4')  # Merge cells for the "No data available" message
            no_data_cell = ws['B4']
            no_data_cell.value = "No data available"
            no_data_cell.alignment = Alignment(horizontal='center')
            no_data_cell.font = Font(bold=True, size=14)
            logging.info("No data available message added.")
        else:
            # Write DataFrame to Excel
            for r_idx, row in enumerate(df.iterrows(), start=4):
                for c_idx, value in enumerate(row[1], start=1):
                    if isinstance(value, pd.Timestamp):
                        value = value.tz_localize(None)
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
            logging.info("Data written to worksheet.")

        # Apply borders to data
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )

        for row in ws.iter_rows(
            min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = border_style
        logging.info("Borders applied to data.")

        # Set column names (headers)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_style
        logging.info("Column headers set.")

        # Set column widths
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            column_letter = get_column_letter(col_idx)  # Convert column index to letter
            ws.column_dimensions[column_letter].width = adjusted_width
        logging.info("Column widths set.")

        # Save the workbook
        wb.save(filename)
        logging.info(f"Workbook saved as {filename}")

    except Exception as e:
        logging.error(f"Error: {str(e)}")
        print(f"Error: {str(e)}")

def generate_pdf_report_2(data, filename, title_text):
    """
    Generate PDF report from provided data.

    Args:
    data (list): List of dictionaries containing the data
    filename (str): Output PDF filename
    title_text (str): Report title
    """
    try:
        cursor_logger = LOGGINGS.CustomLogger()
        logging = cursor_logger.setup_logger()

        global custom_page_size
        time = datetime.now()
        logging.info(f"Report generation started at {time}.")

        # Extract columns from the first dictionary in the data list
        columns = list(data[0].keys())
        logging.info(f"Columns extracted: {columns}")

        batch_size = 1000
        long_text_columns = None
        col_widths = None
        all_tables = []

        # Set page size based on column count
        column_count = len(columns)
        logging.info(f"Column count: {column_count}")

        if column_count <= 5:
            custom_page_size = A4
        else:
            custom_page_size = A1

        if column_count > 12:
            logging.error("Column limit exceeded, max columns allowed = 12")
            return {
                "status": 500,
                "Error": "Column limit exceeded, max columns allowed = 12",
            }

        page_width = custom_page_size[0]
        logging.info(f"Page size set to: {custom_page_size}")

        # Process results in batches
        for batch_count in range(0, len(data), batch_size):
            result_batch = data[batch_count:batch_count + batch_size]
            logging.info(f"Processing batch {batch_count // batch_size + 1}")

            # Detect long text columns and calculate widths
            if long_text_columns is None:
                long_text_columns = detect_long_text_columns(result_batch)
                col_widths = calculate_column_widths_mysql(
                    page_width, len(result_batch[0]), result_batch
                )
                logging.info(f"Long text columns detected: {long_text_columns}")
                logging.info(f"Column widths calculated: {col_widths}")

            # Prepare table data
            table_data = []
            keys = list(result_batch[0].keys())
            table_data.append(tuple(keys))
            logging.info("Table headers added.")

            # Handle long text columns
            font_size = 8
            for col in long_text_columns:
                col_index = keys.index(col)
                max_content_length = 3000
                line_length = int(0.19 * col_widths[col_index])
                for row in result_batch:
                    row[col] = "\n".join(
                        split_text(
                            str(row[col])[:max_content_length],
                            line_length
                        )
                    )
            logging.info("Long text columns processed.")

            # Process rows
            for row in result_batch:
                new_row = list(row.values())
                modified_row = correct_rows(new_row)
                table_data.append(tuple(modified_row))
            logging.info("Rows processed and added to table data.")

            # Create and style table
            table = Table(table_data, colWidths=col_widths)
            style = TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), font_size),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
            table.setStyle(style)
            all_tables.append(table)
            logging.info("Table created and styled.")

        # Generate PDF
        def on_first_page_function(canvas, doc):
            first_page_content(canvas, doc, title_text, time, custom_page_size)
            logging.info("First page function executed.")

        output_doc = SimpleDocTemplate(
            filename, pagesize=(custom_page_size[1], custom_page_size[0])
        )
        output_doc.build(all_tables, onFirstPage=on_first_page_function)
        logging.info(f"PDF generated successfully: {filename}")

    except Exception as e:
        logging.error(f"Error: {str(e)}")
        print(f"Error: {str(e)}")

def generate_csv_report(df, filename):
    """
    Generate a CSV report with basic formatting.
    While CSV cannot store rich formatting like Excel, this function:
    - Adds a title row
    - Properly formats headers
    - Handles null values
    - Ensures consistent formatting
    - Adds metadata like generation time
   
    Args:
        df (pandas.DataFrame): Input DataFrame
        filename (str): Output CSV filename
        title_text (str): Report title
    """
    # Create a copy of the dataframe to avoid modifying the original
    df_copy = df.copy()
   
    # Format column names (replace underscores with spaces and title case)
    df_copy.columns = [col.replace('_', ' ').title() for col in df_copy.columns]
   
    # Create the rows for our CSV
    rows_to_write = []
   
    if df_copy.empty:
        # Handle empty DataFrame
        rows_to_write.append(["No data available"])
    else:
        # Add headers
        rows_to_write.append(df_copy.columns.tolist())
       
        # Add data rows
        for _, row in df_copy.iterrows():
            # Convert None to empty string and format other values
            formatted_row = []
            for value in row:
                if pd.isna(value):
                    formatted_row.append('')
                elif isinstance(value, (int, float)):
                    formatted_row.append(f"{value:,}")  # Add thousand separators
                else:
                    formatted_row.append(str(value))
            rows_to_write.append(formatted_row)
   
    # Write to CSV
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(rows_to_write)
