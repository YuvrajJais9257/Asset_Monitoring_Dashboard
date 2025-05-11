import openpyxl
import re
import os
import csv
import pandas as pd
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as excel_img
from fastapi import FastAPI, Request, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from datetime import datetime
import psycopg2
from reportlab.lib.pagesizes import A1
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import mysql.connector
import vertica_python
from sqlalchemy import create_engine
import zlib
import io
import ast
from psycopg2 import Error


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
FILES_DIR = "/files"
os.makedirs(FILES_DIR, exist_ok=True)
FILES_DIR = os.path.join(os.getcwd(), "files")

# host = 'db'
# port = 3306
# username = 'root'
# password = 'Admin123*'
# schema = 'hyphenview_new_release_test'

mysql_connection = {
    'host' : 'db',
    'port' : 3306,
    'username' : 'root',
    'password' : 'Admin123*',
    'schema' : 'hyphenview_new_release_test'
}

postgres_connection = {
    'host' : 'postgres',
    'port' : 5432,
    'user' : 'postgres',
    'password' : 'Admin123*',
    'dbname' : 'hyphenview_new'
}

# def fetch_image_from_mysql(db_user, db_password, db_host, db_name, table_name, column_name, condition, output_path):
def fetch_image_from_mysql(customer_id):
    """
    Fetches an image stored as a BLOB in a MySQL database and saves it to a file.

    :param db_user: MySQL username
    :param db_password: MySQL password
    :param db_host: MySQL host (e.g., 'localhost')
    :param db_name: Name of the database
    :param table_name: Table containing the image
    :param column_name: Column storing the image as a BLOB
    :param condition: SQL WHERE condition (e.g., "id=1")
    :param output_path: Path where the image should be saved
    :return: Path to the saved image file or None if no image is found
    """

    try:
        # Connect to MySQL database
        conn = mysql.connector.connect(
            host=mysql_connection['host'],
            user=mysql_connection['username'],
            port=mysql_connection['port'],
            password=mysql_connection['password'],
            database=mysql_connection['schema']
        )
        cursor = conn.cursor()

        # Fetch the BLOB data
        query = f"SELECT pdf_logo FROM img WHERE customer_id = %s LIMIT 1"
        cursor.execute(query,customer_id)
        result = cursor.fetchone()

        if result and result[0]:  # Ensure an image was found
            img_data = result[0]

            # Save image to file
            with open('logo.jpg', "wb") as img_file:
                img_file.write(img_data)

            return 'logo.jpg'

        else:
            return None

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return None

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def fetch_image_from_postgres(customer_id):
    """
    Fetches an image stored as a BLOB in a MySQL database and saves it to a file.

    :param db_user: MySQL username
    :param db_password: MySQL password
    :param db_host: MySQL host (e.g., 'localhost')
    :param db_name: Name of the database
    :param table_name: Table containing the image
    :param column_name: Column storing the image as a BLOB
    :param condition: SQL WHERE condition (e.g., "id=1")
    :param output_path: Path where the image should be saved
    :return: Path to the saved image file or None if no image is found
    """

    try:
        # Connect to MySQL database
        conn = psycopg2.connect(
            host=postgres_connection['host'],
            user=postgres_connection['user'],
            port=postgres_connection['port'],
            password=postgres_connection['password'],
            dbname=postgres_connection['dbname']
        )
        cursor = conn.cursor()
        

        # Fetch the BLOB data
        query = f"SELECT pdf_logo FROM img WHERE customer_id = %s"
        cursor.execute(query,(customer_id,))
        result = cursor.fetchone()
        

        if result and result[0]:  # Ensure an image was found
            img_data = result[0]

            if isinstance(img_data, memoryview):
                img_data = img_data.tobytes()

            # Save image to file
            with open('logo.jpg', "wb") as img_file:
                img_file.write(img_data)

            return 'logo.jpg'

        else:
            return None

    except Error as err:
        print(f"Postgres Error: {err}")
        return None

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def html_text_convertor(html):
    try:

        text = re.sub(r'<br\s*/?>', ' ', html)
        text = re.sub(r'<style[^>]*>.*?</style>|<script[^>]*>.*?</script>', '', text)
        text = re.sub(r'<.*?>', '', text)
        text = re.sub(r'&\w+;', ' ', text)
        return text

    except Exception as err:
        raise ValueError(f"Error: {err}")

def correct_rows(row):
    modified_row = []
    for item in row:
        if isinstance(item, str) and ('<' in item or '>' in item):
            new_item = html_text_convertor(item)
            modified_row.append(new_item)
        elif item == None:
            modified_row.append('None')
        elif isinstance(item, list):
            modified_row.append(str(item))
        elif isinstance(item, datetime):
            modified_row.append(item.replace(tzinfo=None))
        else:
            modified_row.append(item)
    return modified_row

def save_uploaded_file(file_content,filename):
    filepath = filename
    with open(filepath, "wb") as f:
        f.write(file_content)
    return(filepath)

def split_text(text, max_length):
    return [text[i:i + max_length] for i in range(0, len(text), max_length)]

def detect_long_text_columns(rows, max_length=10):
    long_text_columns = []
    for key in rows[0].keys():
        for row in rows:
            if isinstance(row[key], str) and len(row[key]) > max_length:
                long_text_columns.append(key)
                break
    return long_text_columns

def detect_long_text_columns_postgres(rows, columns, max_length=10):
    """
    Detect long text columns in PostgreSQL result with explicit column information.
    """
    long_text_columns = []
    keys = [col.name for col in columns]
    for i, key in enumerate(keys):
        for row in rows:
            if isinstance(row[i], str) and len(row[i]) > max_length:
                long_text_columns.append(key)
                break

    return long_text_columns

def first_page_content(db_type,customer_id, user_email_id, canvas, doc, report_title, time, start_date=0, end_date=0):
    """
    Adds an image (e.g., logo) at the top-left, aligns the report title in the same row, 
    and includes the report period below the report generation timestamp.

    :param customer_id: Unique identifier for the customer (not used directly here).
    :param canvas: The ReportLab canvas to draw on.
    :param doc: The document object (not directly used here).
    :param report_title: Title of the report.
    :param time: Timestamp for when the report is generated.
    :param image_path: Path to the image file (e.g., logo) to be placed at the top-left.
    :param start_date: Start date of the report period.
    :param end_date: End date of the report period.
    """
    page_width, page_height = A1  # Get A1 page dimensions
    formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
    
    # Set up image position (top-left)
    img_width = 200  # Adjust as needed
    img_height = 80  # Adjust as needed
    img_x = 50  # Left margin
    img_y = A1[0] - 100  # From the top

    # Draw image (logo)
    try:
        # if db_type == "mysql":
        #     img_path = fetch_image_from_mysql(customer_id)
        # elif db_type == "postgres":
        img_path = fetch_image_from_postgres(customer_id)

        canvas.drawImage(img_path, img_x, img_y, width=img_width, height=img_height, mask='auto')
    except Exception as e:
        pass
        #print(f"Error loading image: {e}")

    canvas.setFont("Helvetica-Bold", 14)
    text_width = canvas.stringWidth(user_email_id, "Helvetica-Bold", 14)
    x_text_below = img_x + (img_width - text_width) / 2  # Centered below the image
    y_text_below = img_y - 20  # Spacing between image and text
    #canvas.drawString(x_text_below, y_text_below, user_email_id)

    # Report Title (aligned horizontally with the image)
    canvas.setFont("Helvetica-Bold", 18)
    title_width = canvas.stringWidth(report_title, "Helvetica-Bold", 18)
    x_title = (page_height - title_width) / 2  # Center the title
    y_title = img_y + (img_height / 2)  # Align title with the middle of the image
    canvas.drawString(x_title, y_title, report_title)

    # Current Time (Right-Aligned)
    current_time = f"Report Generated at: {formatted_datetime}"
    # time_width = canvas.stringWidth(current_time, "Helvetica", 14)
    x_text = A1[1] - 430
    y_text = A1[0] - 30
    canvas.setFont("Helvetica-Bold", 15)
    canvas.drawString(x_text, y_text, current_time)

    # Report Period (Below "Report Generated at")
    if start_date != 0 and end_date != 0:
        time_period = f"Report Period: {start_date} to {end_date}"
        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawString(x_text, y_text-20, time_period)  # Positioned below current time
    else:
        gen_by = f"Report Generated By: {user_email_id}"
        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawString(x_text, y_text-20, gen_by)

def calculate_column_widths_mysql(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content.
    """
    flag = 0
    max_column_widths = [len(col) for col in rows[0].keys()]
    for row in rows:
        for i, key in enumerate(row.keys()):
            max_width = len(str(row[key]))
            max_column_widths[i] = max(max_width, max_column_widths[i])

    total_width = sum(max_column_widths)

    while total_width < page_width:
        scaling_factor = (page_width) / total_width
        max_column_widths = [max(min_absolute_width, (width * scaling_factor)) for width in max_column_widths]
        total_width = sum(max_column_widths)
        flag = 1

    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [max(min_absolute_width, width * scaling_factor) for width in max_column_widths]
    # Ensure that each column has a minimum width of min_absolute_width

    # Adjust font size based on the number of columns
    # max_font_size = 12
    # min_font_size = 8
    # font_size = max(min_font_size, min(max_font_size, 0.5 * (page_width / (num_columns * max_font_size))))

    return column_widths

def calculate_column_widths_postgres(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content for PostgreSQL cursor result.
    """
    flag = 0
    max_column_widths = [len(str(value)) for value in rows[0]]
    for row in rows:
        for i, value in enumerate(row):
            max_width = len(str(value))
            max_column_widths[i] = max(max_width, max_column_widths[i])
    total_width = sum(max_column_widths)
    count = 0
    while total_width < page_width:
        count +=1
        scaling_factor = (page_width) / total_width
        max_column_widths = [max(min_absolute_width, (width * scaling_factor)) for width in max_column_widths]
        total_width = sum(max_column_widths)
        flag = 1

    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [max(min_absolute_width, int(width * scaling_factor)) for width in max_column_widths]

    return column_widths

def fetch_mysql_data(user_email_id, customer_id, host, port, database, user, password, query, template_file_path, report_title, start_date=0, end_date=0):

    try:
        connection = mysql.connector.connect(
            host=host,
            port=port,
            database=database,
            user=user,
            password=password
        )
        time = datetime.now()
        formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
        current_time = "Report Generated at: " + formatted_datetime
        cursor = connection.cursor(dictionary=True)
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        img_path = fetch_image_from_mysql(customer_id)
        img = excel_img(img_path)
        img.width = 100
        img.height = 50
        worksheet.add_image(img, "A1")  # Add image to cell A1
        worksheet.row_dimensions[1].height = 50

        if len(columns) >= 1:
            title_col_range = (
                f"A1:{chr(65 + len(columns))}1"  # Determine column range for title
            )
        else:
            title_col_range = "A1:A1"
        worksheet.merge_cells(title_col_range)  # Merge cells for title
        title_cell = worksheet["A1"]  # Assuming the title starts from column B
        title_cell.value = report_title
        title_cell.font = Font(bold=True, size=18)

        if start_date!=0 and end_date!=0:
            time_period = "Report Period: " + start_date + " to " + end_date
            worksheet.append([time_period])
        else:
            worksheet.append([None])
        worksheet.append([current_time])
        gen_by = f"Generated By : {user_email_id}"
        worksheet.append([gen_by])
        worksheet.append([None])
        worksheet.append(columns)
        for row in cursor.fetchall():
            new_row = list(row.values())
            modified_row = correct_rows(new_row)
            worksheet.append(tuple(modified_row))
        connection.close()
        for row in worksheet.iter_rows(min_row=1, max_row=6):
            for cell in row:
                cell.font = Font(bold=True)
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        valid_filename = f"{template_file_path}_{current_datetime_str}.xlsx"
        workbook.save(valid_filename)
        return {"status":200,"filename":valid_filename}

    except Exception as e:
        return {"status":500,"error":f"Error: {e}"}

def fetch_postgres_data(user_email_id, customer_id, host, port, database, user, password, query, template_file_path, report_title, start_date=0, end_date=0):
    try:
        connection = psycopg2.connect(
            host=host,
            port=port,
            database=database,
            user=user,
            password=password
        )
        time = datetime.now()
        formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
        current_time = "Report Generated at: " + formatted_datetime
        cursor = connection.cursor()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        cursor.execute(query)

        columns = [desc[0] for desc in cursor.description]
        img_path = fetch_image_from_postgres(customer_id)
        img = excel_img(img_path)
        img.width = 100
        img.height = 50
        worksheet.add_image(img, "A1")  # Add image to cell A1
        worksheet.row_dimensions[1].height = 50

        if len(columns) >= 1:
            title_col_range = (
                f"A1:{chr(65 + len(columns))}1"  # Determine column range for title
            )
        else:
            title_col_range = "A1:A1"
        worksheet.merge_cells(title_col_range)  # Merge cells for title
        title_cell = worksheet["A1"]  # Assuming the title starts from column B
        title_cell.value = report_title
        title_cell.font = Font(bold=True, size=18)
        if start_date!=0 and end_date!=0:
            time_period = "Report Period: " + start_date + " to " + end_date
            worksheet.append([time_period])
        else:
            worksheet.append([None])
        worksheet.append([current_time])
        gen_by = f"Generated By : {user_email_id}"
        worksheet.append([gen_by])
        worksheet.append([None])
        worksheet.append(columns)
        for row in cursor.fetchall():
            modified_row = correct_rows(row)
            worksheet.append(tuple(modified_row))
        connection.close()
        for row in worksheet.iter_rows(min_row=1, max_row=6):
            for cell in row:
                cell.font = Font(bold=True)
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        valid_filename = f"{template_file_path}_{current_datetime_str}.xlsx"
        workbook.save(valid_filename)
        return {"status":200,"filename":valid_filename}

    except Exception as e:
        return {"status":500,"error":f"Error: {e}"}

def fetch_vertica_data(user_email_id, customer_id, host, port, database, user, password, query, template_file_path, report_title, start_date=0, end_date=0):
    try:
        conn_info = {
            'host': host,
            'port': port,
            'user': user,
            'password': password,
            'schema': database
            }
        connection = vertica_python.connect(**conn_info)
        time = datetime.now()
        formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
        current_time = "Report Generated at: " + formatted_datetime
        cursor = connection.cursor()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        cursor.execute(query)

        columns = [desc[0] for desc in cursor.description]
        img_path = fetch_image_from_mysql(customer_id)
        img = excel_img(img_path)
        img.width = 100
        img.height = 50
        worksheet.add_image(img, "A1")  # Add image to cell A1
        worksheet.row_dimensions[1].height = 50

        if len(columns) >= 1:
            title_col_range = (
                f"A1:{chr(65 + len(columns))}1"  # Determine column range for title
            )
        else:
            title_col_range = "A1:A1"
        worksheet.merge_cells(title_col_range)  # Merge cells for title
        title_cell = worksheet["A1"]  # Assuming the title starts from column B
        title_cell.value = report_title
        title_cell.font = Font(bold=True, size=18)
        if start_date!=0 and end_date!=0:
            time_period = "Report Period: " + start_date + " to " + end_date
            worksheet.append([time_period])
        else:
            worksheet.append([None])
        worksheet.append([current_time])
        gen_by = f"Generated By : {user_email_id}"
        worksheet.append([gen_by])
        worksheet.append([None])
        worksheet.append(columns)
        for row in cursor.fetchall():
            modified_row = correct_rows(row)
            worksheet.append(tuple(modified_row))
        connection.close()
        for row in worksheet.iter_rows(min_row=1, max_row=6):
            for cell in row:
                cell.font = Font(bold=True)
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        valid_filename = f"{template_file_path}_{current_datetime_str}.xlsx"
        workbook.save(valid_filename)
        return {"status":200,"filename":valid_filename}

    except Exception as e:
        return {"status":500,"error":f"Error: {e}"}

def get_dataframe(dbtype, db_user, db_password, db_host, db_port, db_name, query):
    """Connect to a database dynamically and return a Pandas DataFrame."""

    # Define connection strings for different database types
    if dbtype == "postgres":
        engine = create_engine(f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}")
    elif dbtype == "mysql":
        engine = create_engine(f"mysql+mysqlconnector://{db_user}:{db_password}@{db_host}/{db_name}")
    elif dbtype == "vertica":
        engine = create_engine(f"vertica+vertica_python://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}")
    else:
        raise ValueError("Unsupported database type. Choose from: 'postgres', 'mysql', 'vertica'.")

    # Execute query and return DataFrame
    df = pd.read_sql(query, engine)
    return df

def read_records(connection, table: str, columns: list = None, where_conditions: dict = None
    ):
        """
        Reads records from the specified table with optional column\
              filtering and multiple WHERE conditions.
 
        :param table: Name of the table.
        :param columns: List of column names to retrieve. If None, selects all columns.
        :param where_conditions: Dictionary of column-value pairs for WHERE clause.
        :return: List of retrieved records.
        """
        if not connection:
            print("No active database connection.")
            return
 
        # Prepare SELECT clause
        select_clause = ", ".join(columns) if columns else "*"
 
        # Prepare WHERE clause if conditions are provided
        if where_conditions:
            where_clause = " AND ".join(
                [f"{key} = %s" for key in where_conditions.keys()]
            )
            where_values = tuple(where_conditions.values())
            query = f"SELECT {select_clause} FROM {table} WHERE {where_clause}"
        else:
            query = f"SELECT {select_clause} FROM {table}"
            where_values = {}
 
        try:
            cursor = connection.cursor()
            cursor.execute(query, where_values)
            records = []
            for row in cursor.fetchall():
                records.append(dict(zip([d[0] for d in cursor.description], row)))
            return records
        except Exception as e:
            print(e)

def generate_where_clause_mysql(filter_options, filter_operations):
    where_clauses = []
    # Set up logging
    for filter_option, filter_operation in zip(filter_options, filter_operations):
        for column, values in filter_option.items():
            operation = filter_operation.get(column)

            if values == []:
                continue              

            if operation == "equals":
                clause = f"`{column}` IN ({', '.join([f'{value}' for value in values])})"
            elif operation == "notEquals":
                clause = f"`{column}` NOT IN ({', '.join([f'{value}' for value in values])})"
            elif operation in ("contains", "fuzzy"):
                clause = " OR ".join([f"`{column}` LIKE '%{value}%'" for value in values])
            elif operation == "startsWith":
                clause = " OR ".join([f"`{column}` LIKE '{value}%'" for value in values])
            elif operation == "endsWith":
                clause = " OR ".join([f"`{column}` LIKE '%{value}'" for value in values])
            elif operation == "between":
                if len(values) == 2:
                    clause = f"`{column}` BETWEEN '{values[0]}' AND '{values[1]}'"
                else:
                    raise ValueError("Between inclusive operation requires exactly 2 values.")
            elif operation == "greaterThan":
                clause = f"`{column}` > '{values[0]}'"
            elif operation == "greaterThanOrEqualTo":
                clause = f"`{column}` >= '{values[0]}'"
            elif operation == "lessThan":
                clause = f"`{column}` < '{values[0]}'"
            elif operation == "lessThanOrEqualTo":
                clause = f"`{column}` <= '{values[0]}'"
            else:
                raise ValueError(f"Unsupported operation: {operation}")

            where_clauses.append(f"({clause})")

    return " AND ".join(where_clauses)

def generate_where_clause_postgres(filter_options, filter_operations):
    where_clauses = []
 
    for filter_option, filter_operation in zip(filter_options, filter_operations):
        for column, values in filter_option.items():
            quoted_column = f'"{column}"'
            operation = filter_operation.get(column)
            if values == []:
                continue  
            if operation == "equals":
                formatted_values = [f"'{str(v)}'" for v in values]
                clause = f"{quoted_column} IN ({', '.join(formatted_values)})"
            elif operation == "notEquals":
                formatted_values = [f"'{str(v)}'" for v in values]
                clause = f"{quoted_column} NOT IN ({', '.join(formatted_values)})"
            elif operation in ("contains", "fuzzy"):
                conditions = [f"{quoted_column} ILIKE '%{str(v)}%'" for v in values]
                clause = " OR ".join(conditions)
            elif operation == "startsWith":
                conditions = [f"{quoted_column} ILIKE '{str(v)}%'" for v in values]
                clause = " OR ".join(conditions)
            elif operation == "endsWith":
                conditions = [f"{quoted_column} ILIKE '%{str(v)}'" for v in values]
                clause = " OR ".join(conditions)
            elif operation == "between":
                if len(values) == 2:
                    clause = f"{quoted_column} BETWEEN '{str(values[0])}' AND '{str(values[1])}'"
                else:
                    raise ValueError("Between inclusive operation requires exactly 2 values.")
            elif operation == "greaterThan":
                clause = f"{quoted_column} > '{str(values[0])}'"
            elif operation == "greaterThanOrEqualTo":
                clause = f"{quoted_column} >= '{str(values[0])}'"
            elif operation == "lessThan":
                clause = f"{quoted_column} < '{str(values[0])}'"
            elif operation == "lessThanOrEqualTo":
                clause = f"{quoted_column} <= '{str(values[0])}'"
            else:
                raise ValueError(f"Unsupported operation: {operation}")
 
            where_clauses.append(f"({clause})")
 
    return " AND ".join(where_clauses)

async def generate_report_pdf(user_email_id, customer_id, db_type, query, host, username, password, database, port, report_name, report_title, start_date=None, end_date=None):
    try:
        if port:
            port = port
        else:
            if db_type == 'mysql':
                port = 3306
            elif db_type == 'postgres':
                port = 5432

        time = datetime.now()
        new_formatted_time = time.strftime("%Y-%m-%d_%H-%M-%S")
        # report_name = os.path.join("/root/hyphenview_jan_release_25/hyphen_alchemy_4/app/files", report_name)
        # report_name = os.path.join("/home/abhishek/generate_report_backup/files", report_name)
        report_name = os.path.join(FILES_DIR, report_name)
        

        db_config = {
            'host': host,
            'user': username,
            'password': password,
            'database': database,
            'port' : port
        }

        if db_type == 'mysql':
            conn = mysql.connector.connect(**db_config)

            cursor = conn.cursor(dictionary=True)
            cursor.execute(query)

            batch_size = 1000
            page_width, page_height = A1
            long_text_columns = None
            col_widths = None
            all_tables = []
            column_count = len(cursor.description)
            if column_count <= 12:
                for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
                    if not result_batch:
                        break

                    if long_text_columns is None:
                        long_text_columns = detect_long_text_columns(result_batch)
                        col_widths = calculate_column_widths_mysql(page_width, len(result_batch[0]), result_batch)

                    table_data = []
                    keys = list(result_batch[0].keys())
                    table_data.append(tuple(keys))

                    font_size = 8
                    for col in long_text_columns:
                        col_index = keys.index(col)
                        max_content_length = 3000 # Adjust the max_content_length for cell
                        line_length = int(0.19 * col_widths[col_index])
                        for row in result_batch:
                            row[col] = '\n'.join(split_text(str(row[col])[:max_content_length], line_length))  # Adjust the max_length as needed

                    for row in result_batch:
                        new_row = list(row.values())
                        modified_row = correct_rows(new_row)
                        table_data.append(tuple(modified_row))

                    table = Table(table_data, colWidths=col_widths)
                    style = TableStyle([
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])

                    table.setStyle(style)

                    all_tables.append(table)
            else:
                return{"status":500,"Error":"Column limit exceeded, max columns allowed = 12"}

        elif db_type == 'postgres' or db_type == 'vertica':
            if db_type == 'postgres':
                conn = psycopg2.connect(**db_config)
            elif db_type == 'vertica':
                conn_info = {
                    'host': host,
                    'port': port,
                    'user': username,
                    'password': password,
                    'schema': database
                    }
                conn = vertica_python.connect(**conn_info)

            cursor = conn.cursor()
            cursor.execute(query)

            batch_size = 1000
            page_width, page_height = A1
            long_text_columns = None
            col_widths = None
            all_tables = []
            column_count = len(cursor.description)
            if column_count <= 12:
                for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
                    if not result_batch:
                        break
                    long_text_columns = detect_long_text_columns_postgres(result_batch, cursor.description)
                    col_widths = calculate_column_widths_postgres(page_width, len(result_batch[0]), result_batch)
                    table_data = []
                    font_size = 8

                    keys = [desc[0] for desc in cursor.description]
                    table_data.append(keys)

                    for col in long_text_columns:
                        col_index = keys.index(col)
                        max_content_length = 3000
                        line_length = int(0.19 * col_widths[col_index])
                        result_batch = [list(row) for row in result_batch]

                        for row in result_batch:
                            row[col_index] = '\n'.join(split_text(str(row[col_index])[:max_content_length], line_length))

                    for row in result_batch:
                        modified_row = correct_rows(row)
                        table_data.append(tuple(modified_row))

                    table = Table(table_data, colWidths=col_widths)
                    style = TableStyle([
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])

                    table.setStyle(style)

                    all_tables.append(table)

            else:
                return{"status":500,"Error":"Column limit exceeded, max columns allowed = 12"}

        valid_filename = f"{report_name}_{new_formatted_time}.pdf"
        #report_title = f"{report_title}_{new_formatted_time}.pdf"
        report_title = report_title
        def on_first_page_function(canvas, doc):
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                first_page_content(db_type,customer_id, user_email_id, canvas, doc, report_title, time)
            else:
                first_page_content(db_type,customer_id, user_email_id, canvas, doc, report_title, time, start_date, end_date)

        output_doc = SimpleDocTemplate(valid_filename, pagesize=(A1[1], A1[0]))
        output_doc.build(all_tables, onFirstPage=on_first_page_function)
        cursor.close()
        conn.close()

        print(valid_filename)
        return {"message":"Report Generated Successfully.","file_url":valid_filename,"report_title":report_title}
    except Exception as e:
        return HTTPException(status_code=500, detail=f"Error: {str(e)}")

async def generate_report_excel(user_email_id, customer_id, db_type, query, host, username, password, database, port, report_name, report_title, start_date=None, end_date=None):
    try:
        host = host
        schema = database
        username = username
        password = password
        # report_name = os.path.join("/home/abhishek/generate_report_backup/files", report_name)
        report_name = os.path.join(FILES_DIR, report_name)
        if port:
            port = port
        else:
            if db_type == 'mysql':
                port = 3306
            elif db_type == 'postgres':
                port = 5432

        if db_type == 'mysql':
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                start_date=0
                end_date=0
            file = fetch_mysql_data(user_email_id, customer_id, host, port, schema, username, password, query, report_name, report_title, start_date, end_date)
            if file["status"] == 500:
                return file
            else:
                report_title = file["filename"].strip().split('/')[-1]
                return {"message":"Report Generated Successfully.","file_url":file["filename"],"report_title":report_title}

        elif db_type == 'postgres':
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                start_date=0
                end_date=0
            file = fetch_postgres_data(user_email_id, customer_id, host, port, schema, username, password, query, report_name, report_title, start_date, end_date)
            if file["status"] == 500:
                return file
            else:
                report_title = file["filename"].strip().split('/')[-1]
                return {"message":"Report Generated Successfully.","file_url":file["filename"],"report_title":report_title}

        elif db_type == 'vertica':
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                start_date=0
                end_date=0
            file = fetch_vertica_data(user_email_id, customer_id, host, port, schema, username, password, query, report_name, report_title, start_date, end_date)
            if file["status"] == 500:
                return file
            else:
                report_title = file["filename"].strip().split('/')[-1]
                return {"message":"Report Generated Successfully.","file_url":file["filename"],"report_title":report_title}
    except Exception as e:
        return HTTPException(status_code=500, detail=f"Error: {str(e)}")

async def generate_report_csv(user_email_id, db_type, query, host, username, password, database, port, report_name, report_title, start_date=None, end_date=None):
    """
    Generate a CSV report with basic formatting.
    While CSV cannot store rich formatting like Excel, this function:
    - Adds a title row
    - Properly formats headers
    - Handles null values
    - Ensures consistent formatting
    - Adds metadata like generation time
   
    Args:
        df (pandas.DataFrame): Input DataFrame
        filename (str): Output CSV filename
        title_text (str): Report title
    """
    try:
        time = datetime.now()
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        # report_name = os.path.join("/home/abhishek/generate_report_backup/files", report_name)
        report_name = os.path.join(FILES_DIR, report_name)
        valid_filename = f"{report_name}_{current_datetime_str}.csv"
        # Create a copy of the dataframe to avoid modifying the original
        df = get_dataframe(db_type, username, password, host, port, database, query)
        df_copy = df.copy()
    
        # Format column names (replace underscores with spaces and title case)
        df_copy.columns = [col.replace('_', ' ').title() for col in df_copy.columns]
    
        # Create the rows for our CSV
        rows_to_write = []
        rows_to_write.append([f"**{report_title}**"])
        if start_date!= None and end_date!= None:
            time_period = "Report Period: " + start_date + " to " + end_date
            rows_to_write.append([f"**{time_period}**"])
        rows_to_write.append([f"**Report Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}**"])
        gen_by = f"**Generated By : {user_email_id}**"
        rows_to_write.append([gen_by])
        rows_to_write.append([])
        if df_copy.empty:
            # Handle empty DataFrame
            rows_to_write.append(["No data available"])
        else:
            # Add headers
            rows_to_write.append(df_copy.columns.tolist())
        
            # Add data rows
            for _, row in df_copy.iterrows():
                # Convert None to empty string and format other values
                formatted_row = []
                for value in row:
                    if pd.isna(value):
                        formatted_row.append('')
                    elif isinstance(value, (int, float)):
                        formatted_row.append(f"{value:,}")  # Add thousand separators
                    else:
                        formatted_row.append(str(value))
                rows_to_write.append(formatted_row)
    
        # Write to CSV
        with open(valid_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(rows_to_write)

        report_title = valid_filename.strip().split('/')[-1]
        return {"message":"Report Generated Successfully.","file_url":valid_filename,"report_title":report_title}

    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Error: {str(err)}")

@app.post("/generate_report")
async def extract(user_details: Request):
    try:
        user_details = await user_details.json()

        if "user_email_id" in user_details:
            user_email_id = user_details.get("user_email_id")
        customer_id = user_details.get("customer_id")
        database_type = user_details.get("database_type")
        detailed_report_id = user_details.get("report_id")
        master_report_name = user_details.get("report_name")
        filter_options = user_details.get("filter_options")
        filter_operations = user_details.get("filter_operations")
        file_format = user_details.get("file_format")
        column_names = user_details.get("column_names")
        filter_value = user_details.get("filter_value")
        selectedSeriesName = user_details.get("selectedSeriesName")
        flag = user_details.get("flag")
        sorting_options = user_details.get("sorting_options")

        if not customer_id or not database_type:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content="Missing required parameters: customer_id or database_type.",
            )
        # Select database service
        # if database_type == "mysql":
        #     database_url = {
        #         "host": host,
        #         "port": port,
        #         "username": username,
        #         "password": password,
        #         "database": schema,
        #     }
        # else:
        #     return JSONResponse(
        #         status_code=status.HTTP_400_BAD_REQUEST,
        #         content=f"Unsupported database type: {database_type}",
        #     )

        # Connect to the database
        try:
            if database_type == "postgres":
                conn = psycopg2.connect(
                    host=postgres_connection["host"],
                    port=postgres_connection["port"],
                    user=postgres_connection["user"],
                    password=postgres_connection["password"],
                    dbname=postgres_connection["dbname"],
                )
            elif database_type == "mysql":
                conn = mysql.connector.connect(
                    host=mysql_connection["host"],
                    port=mysql_connection["port"],
                    user=mysql_connection["username"],
                    password=mysql_connection["password"],
                    database=mysql_connection["database"],
                )
            else:
                return JSONResponse(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    content=f"Unsupported database type: {database_type}",
                )
        except Exception as db_connect_error:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content=f"Database connection failed: {db_connect_error}",
            )
        # Fetch master report details
        try:
            if flag != "drilldownpage":
                report_name = master_report_name
                pass
            else:
                if detailed_report_id:
                    where_conditions = {
                        "customer_id": customer_id,
                        "detailed_report_id": detailed_report_id
                    }
                else:
                    where_conditions = {
                        "customer_id": customer_id,
                        "master_report": master_report_name
                    }
                columns = ["*"]
                
                result = read_records(
                    conn, table="detailed_report", columns=columns, where_conditions=where_conditions
                )
                if not result:
                    return JSONResponse(
                        status_code=status.HTTP_404_NOT_FOUND,
                        content="No master_report found for the given details.",
                    )

                report_name = result[0]["drilldown_report"]
                drilldown_column = ast.literal_eval(result[0]["drilldown_column"])
        except Exception as fetch_template_error:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content=f"Error fetching report template: {fetch_template_error}",
            )

        # Fetch report template details
        try:
            print('................',report_name)
            where_conditions = {
                "customer_id": customer_id,
                "report_template_name": report_name
            }
            columns = ["report_type", "defined_query", "db_details_id"]
            
            result = read_records(
                conn, table="report_template", columns=columns, where_conditions=where_conditions
            )
            if not result:
                return JSONResponse(
                    status_code=status.HTTP_404_NOT_FOUND,
                    content="No report templates found for the given details.",
                )

            db_details_id = result[0]["db_details_id"]
            query = result[0]["defined_query"]
        except Exception as fetch_template_error:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content=f"Error fetching report template: {fetch_template_error}",
            )
            

        # Fetch secondary database details
        try:
            columns = [
                "rdbms_name",
                "domain_name",
                "db_port",
                "db_user_name",
                "db_password",
                "db_schema_name",
            ]
            db_details = read_records(
                conn, table='database_details', columns=columns, where_conditions={"db_details_id": db_details_id}
            )
            if not db_details:
                return JSONResponse(
                    status_code=status.HTTP_404_NOT_FOUND,
                    content="No database details found for the given report.",
                )

            db_config = db_details[0]
        except Exception as fetch_db_error:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content=f"Error fetching database details: {fetch_db_error}",
            )
        finally:
            conn.close()

        # Construct WHERE clause based on filter options
        try:    
            query = query.replace(";","")
            if flag == "drilldownpage":
                if db_config["rdbms_name"] == "mysql":
                    if filter_value == "":
                        query = query
                    else:
                        if len(drilldown_column) == 1:
                            condition = (
                                f" WHERE `{drilldown_column[0]}` = '{filter_value}'"
                            )
                        elif len(drilldown_column) == 2:
                            condition = f" WHERE `{drilldown_column[0]}` = '{filter_value}'\
                                    AND `{drilldown_column[1]}` = '{selectedSeriesName}'"
                        query = f"Select * from ({query}) mains" + condition

                elif db_config["rdbms_name"] == 'postgres':
                    if filter_value == "":
                        query = query
                    else:
                        filter_value = filter_value.replace("'","''")
                        selectedSeriesName = selectedSeriesName.replace("'","''")
                        if len(drilldown_column) == 1:
                            condition = f" WHERE \"{drilldown_column[0]}\" = '{filter_value}'"
                        elif len(drilldown_column) == 2:
                            condition = f" WHERE \"{drilldown_column[0]}\" = '{filter_value}'\
                                AND \"{drilldown_column[1]}\" = '{selectedSeriesName}'"
                        query = f"Select * from ({query}) mains" + condition


            if filter_operations and filter_options:
                if db_config["rdbms_name"] == 'mysql':
                    where_clause = generate_where_clause_mysql(filter_options, filter_operations)
                    if column_names and len(column_names) > 0:
                        joined_column_names = ", ".join([f"`{col}`" for col in column_names])
                        query = f'SELECT {joined_column_names} FROM ({query}) AS MAIN WHERE {where_clause}'
                    else:
                        query = f'SELECT * FROM ({query}) AS MAIN WHERE {where_clause}'
                    print(query)
                elif db_config["rdbms_name"] == 'postgres':
                    where_clause = generate_where_clause_postgres(filter_options, filter_operations)
                    if column_names and len(column_names) > 0:
                        joined_column_names = ", ".join([f'"{col}"' for col in column_names])
                        query = f'SELECT {joined_column_names} FROM ({query}) AS MAIN WHERE {where_clause}'
                    else:
                        query = f'SELECT * FROM ({query}) AS MAIN WHERE {where_clause}'
            elif column_names and len(column_names) > 0:
                if db_config["rdbms_name"] == 'postgres':
                    joined_column_names = ", ".join([f'"{col}"' for col in column_names])
                else:
                    joined_column_names = ", ".join([f"`{col}`" for col in column_names])
                query = f'SELECT {joined_column_names} FROM ({query}) AS MAIN'
            else:
                query = f'SELECT * FROM ({query}) AS MAIN'

            # Construct ORDER BY clause based on sorting options
            if sorting_options:
                if db_config["rdbms_name"] == 'mysql':
                    order_by_clause = " ORDER BY " + ", ".join(
                        [f"`{sort['id']}` {'DESC' if sort['desc'] else 'ASC'}" for sort in sorting_options]
                    )
                elif db_config["rdbms_name"] == 'postgres':
                    order_by_clause = " ORDER BY " + ", ".join(
                        [f'"{sort["id"]}" {"DESC" if sort["desc"] else "ASC"}' for sort in sorting_options]
                    )
                elif db_config["rdbms_name"] == 'oracle':
                    order_by_clause = " ORDER BY " + ", ".join(
                        [f'"{sort["id"]}" {"DESC" if sort["desc"] else "ASC"}' for sort in sorting_options]
                    )
                query += order_by_clause

        except Exception as construct_query_error:
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content=f"Error constructing query: {construct_query_error}",
            )
        
        if file_format == "pdf":
            res = await generate_report_pdf(user_email_id, customer_id, db_config['rdbms_name'],query,db_config['domain_name'],db_config['db_user_name'],
                                      db_config['db_password'],db_config['db_schema_name'],db_config['db_port'],master_report_name,master_report_name)
        elif file_format == "excel":
            res = await generate_report_excel(user_email_id,customer_id, db_config['rdbms_name'],query,db_config['domain_name'],db_config['db_user_name'],
                                      db_config['db_password'],db_config['db_schema_name'],db_config['db_port'],master_report_name,master_report_name)
        elif file_format == "csv":
            res = await generate_report_csv(user_email_id, db_config['rdbms_name'],query,db_config['domain_name'],db_config['db_user_name'],
                                      db_config['db_password'],db_config['db_schema_name'],db_config['db_port'],master_report_name,master_report_name)
        print(res)
        file_path = res["file_url"]
        report_title = res["report_title"]
        _, extension = os.path.splitext(file_path)
        
        print(extension)
        
        if report_title.endswith(extension):
            report_title = report_title[: -len(extension)]

        with open(file_path,'rb') as new:
            file_content = new.read()
        compressed_data = zlib.compress(file_content, level=zlib.Z_BEST_COMPRESSION)
        compressed_stream = io.BytesIO(compressed_data)

        def iterfile():
            yield from compressed_stream
        # return StreamingResponse(iterfile(), media_type='application/octet-stream', headers={
        #     'Content-Disposition': f'attachment; filename={report_title}.zip',
        #     'X-Filename':report_title,
        #     'Access-Control-Expose-Headers':'Content-Disposition, X-Filename'})
        return StreamingResponse(
            iterfile(),
            media_type='application/octet-stream',
            headers={
                'Content-Disposition': f'attachment; filename="{report_title}{extension}.zlib"',
                'X-Filename': f'{report_title}{extension}',
                'Access-Control-Expose-Headers': 'Content-Disposition, X-Filename'
            }
        )
    except Exception as err:
        print(err)
###############################################################################################################################
