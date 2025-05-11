import openpyxl
import re
import os
import csv
import pandas as pd
from openpyxl.styles import Font
from fastapi import FastAPI, File, UploadFile, Form, Query, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from datetime import datetime
import psycopg2
from reportlab.lib.pagesizes import A1
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import mysql.connector
import vertica_python
from sqlalchemy import create_engine
import zlib
import io


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
FILES_DIR = "/app/files"
os.makedirs(FILES_DIR, exist_ok=True)
app.mount("/files",StaticFiles(directory=FILES_DIR),name="files")

def html_text_convertor(html):
    try:

        text = re.sub(r'<br\s*/?>', ' ', html)
        text = re.sub(r'<style[^>]*>.*?</style>|<script[^>]*>.*?</script>', '', text)
        text = re.sub(r'<.*?>', '', text)
        text = re.sub(r'&\w+;', ' ', text)
        return text

    except Exception as err:
        raise ValueError(f"Error: {err}")

def correct_rows(row):
    modified_row = []
    for item in row:
        if isinstance(item, str) and ('<' in item or '>' in item):
            new_item = html_text_convertor(item)
            modified_row.append(new_item)
        elif item == None:
            modified_row.append('None')
        elif isinstance(item, list):
            modified_row.append(str(item))
        elif isinstance(item, datetime):
            modified_row.append(item.replace(tzinfo=None))
        else:
            modified_row.append(item)
    return modified_row

def save_uploaded_file(file_content,filename):
    filepath = filename
    with open(filepath, "wb") as f:
        f.write(file_content)
    return(filepath)

def split_text(text, max_length):
    return [text[i:i + max_length] for i in range(0, len(text), max_length)]

def detect_long_text_columns(rows, max_length=10):
    long_text_columns = []
    for key in rows[0].keys():
        for row in rows:
            if isinstance(row[key], str) and len(row[key]) > max_length:
                long_text_columns.append(key)
                break
    return long_text_columns

def detect_long_text_columns_postgres(rows, columns, max_length=10):
    """
    Detect long text columns in PostgreSQL result with explicit column information.
    """
    long_text_columns = []
    keys = [col.name for col in columns]
    for i, key in enumerate(keys):
        for row in rows:
            if isinstance(row[i], str) and len(row[i]) > max_length:
                long_text_columns.append(key)
                break

    return long_text_columns

def first_page_content(canvas, doc, report_title, time, start_date=0, end_date=0):
    # page_width, page_height = A1

    # title_width = canvas.stringWidth(report_title, "Helvetica-Bold", 10)/2
    # x_text = (page_width + title_width) / 2
    formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
    # start_date_str = start_date.strftime("%Y-%m-%d %H:%M:%S")
    x_text = 100
    y_text = A1[0] - 40
    y_new = A1[0] - 60
    canvas.setFont("Helvetica-Bold", 15)
    canvas.drawString(x_text, y_text, report_title)
    if start_date!=0 and end_date!=0:
        time_period = "Report Period: " + start_date + " to " + end_date
        canvas.drawString(x_text, y_new, time_period)
    
    x_text = A1[1] - 400
    canvas.setFont("Helvetica-Bold", 15)
    current_time = "Report Generated at: " + formatted_datetime
    canvas.drawString(x_text, y_text, current_time)

def calculate_column_widths_mysql(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content.
    """
    flag = 0
    max_column_widths = [len(col) for col in rows[0].keys()]
    for row in rows:
        for i, key in enumerate(row.keys()):
            max_width = len(str(row[key]))
            max_column_widths[i] = max(max_width, max_column_widths[i])

    total_width = sum(max_column_widths)

    while total_width < page_width:
        scaling_factor = (page_width) / total_width
        max_column_widths = [max(min_absolute_width, (width * scaling_factor)) for width in max_column_widths]
        total_width = sum(max_column_widths)
        flag = 1

    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [max(min_absolute_width, width * scaling_factor) for width in max_column_widths]
    # Ensure that each column has a minimum width of min_absolute_width

    # Adjust font size based on the number of columns
    # max_font_size = 12
    # min_font_size = 8
    # font_size = max(min_font_size, min(max_font_size, 0.5 * (page_width / (num_columns * max_font_size))))

    return column_widths

def calculate_column_widths_postgres(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content for PostgreSQL cursor result.
    """
    flag = 0
    max_column_widths = [len(str(value)) for value in rows[0]]
    for row in rows:
        for i, value in enumerate(row):
            max_width = len(str(value))
            max_column_widths[i] = max(max_width, max_column_widths[i])
    total_width = sum(max_column_widths)
    count = 0
    while total_width < page_width:
        count +=1
        scaling_factor = (page_width) / total_width
        max_column_widths = [max(min_absolute_width, (width * scaling_factor)) for width in max_column_widths]
        total_width = sum(max_column_widths)
        flag = 1

    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [max(min_absolute_width, int(width * scaling_factor)) for width in max_column_widths]

    return column_widths

def fetch_mysql_data(host, port, database, user, password, query, template_file_path, report_title, start_date=0, end_date=0):

    try:
        connection = mysql.connector.connect(
            host=host,
            port=port,
            database=database,
            user=user,
            password=password
        )
        time = datetime.now()
        formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
        current_time = "Report Generated at: " + formatted_datetime
        cursor = connection.cursor(dictionary=True)
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append([report_title])
        if start_date!=0 and end_date!=0:
            time_period = "Report Period: " + start_date + " to " + end_date
            worksheet.append([time_period])
        worksheet.append([current_time])
        worksheet.append([None])
        worksheet.append(columns)
        for row in cursor.fetchall():
            new_row = list(row.values())
            modified_row = correct_rows(new_row)
            worksheet.append(tuple(modified_row))
            # worksheet.append(list(row.values()))

        connection.close()
        for row in worksheet.iter_rows(min_row=1, max_row=4):
            for cell in row:
                cell.font = Font(bold=True)
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        valid_filename = f"{template_file_path}_{current_datetime_str}.xlsx"
        workbook.save(valid_filename)
        return {"status":200,"filename":valid_filename}

    except Exception as e:
        return {"status":500,"error":f"Error: {e}"}

def fetch_postgres_data(host, port, database, user, password, query, template_file_path, report_title, start_date=0, end_date=0):
    try:
        connection = psycopg2.connect(
            host=host,
            port=port,
            database=database,
            user=user,
            password=password
        )
        time = datetime.now()
        formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
        current_time = "Report Generated at: " + formatted_datetime
        cursor = connection.cursor()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        cursor.execute(query)

        columns = [desc[0] for desc in cursor.description]
        worksheet.append([report_title])
        if start_date!=0 and end_date!=0:
            time_period = "Report Period: " + start_date + " to " + end_date
            worksheet.append([time_period])
        worksheet.append([current_time])
        worksheet.append([None])
        worksheet.append(columns)
        for row in cursor.fetchall():
            modified_row = correct_rows(row)
            worksheet.append(tuple(modified_row))
        connection.close()
        for row in worksheet.iter_rows(min_row=1, max_row=4):
            for cell in row:
                cell.font = Font(bold=True)
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        valid_filename = f"{template_file_path}_{current_datetime_str}.xlsx"
        workbook.save(valid_filename)
        return {"status":200,"filename":valid_filename}

    except Exception as e:
        return {"status":500,"error":f"Error: {e}"}

def fetch_vertica_data(host, port, database, user, password, query, template_file_path, report_title, start_date=0, end_date=0):
    try:
        conn_info = {
            'host': host,
            'port': port,
            'user': user,
            'password': password,
            'schema': database
            }
        connection = vertica_python.connect(**conn_info)
        time = datetime.now()
        formatted_datetime = time.strftime("%Y-%m-%d %H:%M:%S")
        current_time = "Report Generated at: " + formatted_datetime
        cursor = connection.cursor()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        cursor.execute(query)

        columns = [desc[0] for desc in cursor.description]
        worksheet.append([report_title])
        if start_date!=0 and end_date!=0:
            time_period = "Report Period: " + start_date + " to " + end_date
            worksheet.append([time_period])
        worksheet.append([current_time])
        worksheet.append([None])
        worksheet.append(columns)
        for row in cursor.fetchall():
            modified_row = correct_rows(row)
            worksheet.append(tuple(modified_row))
        connection.close()
        for row in worksheet.iter_rows(min_row=1, max_row=4):
            for cell in row:
                cell.font = Font(bold=True)
        current_datetime_str = time.strftime("%Y-%m-%d_%H-%M-%S")
        valid_filename = f"{template_file_path}_{current_datetime_str}.xlsx"
        workbook.save(valid_filename)
        return {"status":200,"filename":valid_filename}

    except Exception as e:
        return {"status":500,"error":f"Error: {e}"}

def get_dataframe(dbtype, db_user, db_password, db_host, db_port, db_name, query):
    """Connect to a database dynamically and return a Pandas DataFrame."""

    # Define connection strings for different database types
    if dbtype == "postgres":
        engine = create_engine(f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}")
    elif dbtype == "mysql":
        engine = create_engine(f"mysql+mysqlconnector://{db_user}:{db_password}@{db_host}/{db_name}")
    elif dbtype == "vertica":
        engine = create_engine(f"vertica+vertica_python://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}")
    else:
        raise ValueError("Unsupported database type. Choose from: 'postgres', 'mysql', 'vertica'.")

    # Execute query and return DataFrame
    df = pd.read_sql(query, engine)
    return df

@app.post("/generate_report_pdf")
async def generate_report(db_type: str = Form(...), query: str = Form(...), host: str = Form(...), username: str = Form(...), password: str = Form(...), database: str = Form(...), port: str = Form(None), report_name : str = Form(...), report_title: str = Form(...), start_date: str = Form(None), end_date: str = Form(None)):
    try:
        if port:
            port = port
        else:
            if db_type == 'mysql':
                port = 3306
            elif db_type == 'postgres':
                port = 5432

        time = datetime.now()
        new_formatted_time = time.strftime("%Y-%m-%d_%H-%M-%S")
        # report_name = os.path.join("/root/hyphenview_jan_release_25/hyphen_alchemy_4/app/files", report_name)
        report_name = os.path.join("/app/files", report_name)

        db_config = {
            'host': host,
            'user': username,
            'password': password,
            'database': database,
            'port' : port
        }

        if db_type == 'mysql':
            conn = mysql.connector.connect(**db_config)

            cursor = conn.cursor(dictionary=True)
            cursor.execute(query)

            batch_size = 1000
            page_width, page_height = A1
            long_text_columns = None
            col_widths = None
            all_tables = []
            column_count = len(cursor.description)
            if column_count <= 12:
                for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
                    if not result_batch:
                        break

                    if long_text_columns is None:
                        long_text_columns = detect_long_text_columns(result_batch)
                        col_widths = calculate_column_widths_mysql(page_width, len(result_batch[0]), result_batch)

                    table_data = []
                    keys = list(result_batch[0].keys())
                    table_data.append(tuple(keys))

                    font_size = 8
                    for col in long_text_columns:
                        col_index = keys.index(col)
                        max_content_length = 3000 # Adjust the max_content_length for cell
                        line_length = int(0.19 * col_widths[col_index])
                        for row in result_batch:
                            row[col] = '\n'.join(split_text(str(row[col])[:max_content_length], line_length))  # Adjust the max_length as needed

                    for row in result_batch:
                        new_row = list(row.values())
                        modified_row = correct_rows(new_row)
                        table_data.append(tuple(modified_row))

                    table = Table(table_data, colWidths=col_widths)
                    style = TableStyle([
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])

                    table.setStyle(style)

                    all_tables.append(table)
            else:
                return{"status":500,"Error":"Column limit exceeded, max columns allowed = 12"}

        elif db_type == 'postgres' or db_type == 'vertica':
            if db_type == 'postgres':
                conn = psycopg2.connect(**db_config)
            elif db_type == 'vertica':
                conn_info = {
                    'host': host,
                    'port': port,
                    'user': username,
                    'password': password,
                    'schema': database
                    }
                conn = vertica_python.connect(**conn_info)

            cursor = conn.cursor()
            cursor.execute(query)

            batch_size = 1000
            page_width, page_height = A1
            long_text_columns = None
            col_widths = None
            all_tables = []
            column_count = len(cursor.description)
            if column_count <= 12:
                for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
                    if not result_batch:
                        break
                    long_text_columns = detect_long_text_columns_postgres(result_batch, cursor.description)
                    col_widths = calculate_column_widths_postgres(page_width, len(result_batch[0]), result_batch)
                    table_data = []
                    font_size = 8

                    keys = [desc[0] for desc in cursor.description]
                    table_data.append(keys)

                    for col in long_text_columns:
                        col_index = keys.index(col)
                        max_content_length = 3000
                        line_length = int(0.19 * col_widths[col_index])
                        result_batch = [list(row) for row in result_batch]

                        for row in result_batch:
                            row[col_index] = '\n'.join(split_text(str(row[col_index])[:max_content_length], line_length))

                    for row in result_batch:
                        modified_row = correct_rows(row)
                        table_data.append(tuple(modified_row))

                    table = Table(table_data, colWidths=col_widths)
                    style = TableStyle([
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])

                    table.setStyle(style)

                    all_tables.append(table)

            else:
                return{"status":500,"Error":"Column limit exceeded, max columns allowed = 12"}

        valid_filename = f"{report_name}_{new_formatted_time}.pdf"
        def on_first_page_function(canvas, doc):
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                first_page_content(canvas, doc, report_title, time)
            else:
                first_page_content(canvas, doc, report_title, time, start_date, end_date)

        output_doc = SimpleDocTemplate(valid_filename, pagesize=(A1[1], A1[0]))
        output_doc.build(all_tables, onFirstPage=on_first_page_function)
        cursor.close()
        conn.close()
        
        # with open(valid_filename, 'rb') as file:
        #     file_data = file.read()
 
        # # Compress the file data
        # compressed_data = zlib.compress(file_data, level=zlib.Z_BEST_COMPRESSION)
 
        # # Create a BytesIO stream from the compressed data
        # compressed_stream = io.BytesIO(compressed_data)
 
        # # Define a generator to stream the data
        # def iterfile():
        #     yield from compressed_stream
        # # Return the streaming response
        # return StreamingResponse(iterfile(), media_type='application/octet-stream', headers={
        #     'Content-Disposition': f'attachment; filename={valid_filename}.zip',
        #     "X-Filename": valid_filename,
        #     "Access-Control-Expose-Headers": "Content-Disposition, X-Filename",
        # })
        # return {"message": f"Report generated successfully.","file":"{}".format(valid_filename)}
        print(valid_filename)
        file_url = f"http://exportapi:9002/files/{valid_filename}"
        return JSONResponse(content={"message":"Report Generated Successfully.","file_url":file_url})
        #response_data = {
        #    "message": "Report generated successfully.",
        #    "file": valid_filename
        #}
        #return JSONResponse(content=response_data)
    except Exception as e:
        return HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/generate_report_excel")
async def extract(db_type:str = Form(...), report_name: str = Form(...), query: str = Form(...), host: str = Form(...), username: str = Form(...), password: str = Form(...), database: str = Form(...), port: str = Form(None), report_title: str = Form(...), start_date: str = Form(None), end_date: str = Form(None)):
    try:
        host = host
        schema = database
        username = username
        password = password
        report_name = os.path.join("/app/files", report_name)
        if port:
            port = port
        else:
            if db_type == 'mysql':
                port = 3306
            elif db_type == 'postgres':
                port = 5432

        if db_type == 'mysql':
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                start_date=0
                end_date=0
            file = fetch_mysql_data(host, port, schema, username, password, query, report_name, report_title, start_date, end_date)
            if file["status"] == 500:
                return file
            else:
                return {"message": f"Report generated successfully.","file":"{}".format(file["filename"])}

        elif db_type == 'postgres':
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                start_date=0
                end_date=0
            file = fetch_postgres_data(host, port, schema, username, password, query, report_name, report_title, start_date, end_date)
            if file["status"] == 500:
                return file
            else:
                return {"message": f"Report generated successfully.","file":"{}".format(file["filename"])}

        elif db_type == 'vertica':
            if (start_date==None and end_date==None) or (start_date=='null' and end_date=='null') or (start_date=='' and end_date==''):
                start_date=0
                end_date=0
            file = fetch_vertica_data(host, port, schema, username, password, query, report_name, report_title, start_date, end_date)
            if file["status"] == 500:
                return file
            else:
                return {"message": f"Report generated successfully.","file":"{}".format(file["filename"])}
    except Exception as e:
        return HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.post("/generate_report_csv")
async def generate_csv_report(db_type:str = Form(...), report_name: str = Form(...), query: str = Form(...), host: str = Form(...), username: str = Form(...), password: str = Form(...), database: str = Form(...), port: str = Form(None), report_title: str = Form(...), start_date: str = Form(None), end_date: str = Form(None)):
    """
    Generate a CSV report with basic formatting.
    While CSV cannot store rich formatting like Excel, this function:
    - Adds a title row
    - Properly formats headers
    - Handles null values
    - Ensures consistent formatting
    - Adds metadata like generation time
   
    Args:
        df (pandas.DataFrame): Input DataFrame
        filename (str): Output CSV filename
        title_text (str): Report title
    """
    time = datetime.now()
    new_formatted_time = time.strftime("%Y-%m-%d_%H-%M-%S")
    report_name = os.path.join("/app/files", report_name)
    # Create a copy of the dataframe to avoid modifying the original
    df = get_dataframe(db_type, username, password, host, port, database, query)
    df_copy = df.copy()
   
    # Format column names (replace underscores with spaces and title case)
    df_copy.columns = [col.replace('_', ' ').title() for col in df_copy.columns]
   
    # Create the rows for our CSV
    rows_to_write = []
    rows_to_write.append([report_title])  
    rows_to_write.append([f"Report Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    rows_to_write.append([])
    if df_copy.empty:
        # Handle empty DataFrame
        rows_to_write.append(["No data available"])
    else:
        # Add headers
        rows_to_write.append(df_copy.columns.tolist())
       
        # Add data rows
        for _, row in df_copy.iterrows():
            # Convert None to empty string and format other values
            formatted_row = []
            for value in row:
                if pd.isna(value):
                    formatted_row.append('')
                elif isinstance(value, (int, float)):
                    formatted_row.append(f"{value:,}")  # Add thousand separators
                else:
                    formatted_row.append(str(value))
            rows_to_write.append(formatted_row)
   
    # Write to CSV
    with open(report_name, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(rows_to_write)

###############################################################################################################################
